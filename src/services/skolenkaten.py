"""Parser for Skolenkäten Excel files.

This module provides functionality to parse Skolenkäten survey data from
Excel files downloaded from Skolinspektionen's website.
"""

import logging
import re
from pathlib import Path
from typing import Optional

from .models import SkolenkatResult, SkolenkatSummary

logger = logging.getLogger(__name__)

# Column indices for the main data (based on Excel structure)
# These are 0-indexed positions in the Excel row
COL_ORG_NUMMER = 0
COL_HUVUDMAN = 1
COL_KOMMUN = 2
COL_SKOLENHETSKOD = 3
COL_SKOLENHET = 4
COL_ANTAL_I_GRUPPEN = 5
COL_ANTAL_SVAR = 6
COL_SVARSFREKVENS = 7

# Index column positions (start after metadata columns)
# The index values appear at specific positions in the row
# These are the correct positions based on actual Excel structure analysis
INDEX_POSITIONS = {
    "information": 8,  # 1. Information om utbildningen
    "stimulans": 42,  # 2. Stimulans
    "stod": 91,  # 3. Stöd
    "kritiskt_tankande": 142,  # 4. Kritiskt tänkande
    "bemotande_larare": 165,  # 5. Bemötande - lärare
    "bemotande_elever": 190,  # 6. Bemötande - elever
    "inflytande": 213,  # 7. Inflytande
    "studiero": 236,  # 8. Studiero
    "trygghet": 263,  # 9. Trygghet
    "forhindra_krankningar": 312,  # 10. Förhindra kränkningar
    "elevhalsa": 337,  # 11. Elevhälsa
    "nojdhet": 360,  # 12. Övergripande nöjdhet (not an index)
}


def parse_respondent_type(filename: str) -> tuple[str, str | None]:
    """Extract respondent type and school form from filename.

    Args:
        filename: Name of the Excel file

    Returns:
        Tuple of (respondent_type, skolform)
    """
    filename_lower = filename.lower()

    # Determine respondent type
    if "elever-grundskola-ak-5" in filename_lower or "elever-ak-5" in filename_lower:
        return "elever-grundskola-ak-5", "grundskola"
    elif "elever-grundskola-ak-8" in filename_lower or "elever-ak-8" in filename_lower:
        return "elever-grundskola-ak-8", "grundskola"
    elif "elever-gymnasieskola-ar-2" in filename_lower or "elever-ar-2" in filename_lower:
        return "elever-gymnasieskola-ar-2", "gymnasieskola"
    elif "larare-grundskola" in filename_lower:
        return "larare-grundskola", "grundskola"
    elif "larare-gymnasieskola" in filename_lower:
        return "larare-gymnasieskola", "gymnasieskola"
    elif "pedagogisk-personal-gymnasieskola" in filename_lower:
        return "larare-gymnasieskola", "gymnasieskola"
    elif "pedagogisk-personal-grundskola" in filename_lower:
        return "larare-grundskola", "grundskola"
    elif (
        "vardnadshavare-forskoleklass" in filename_lower
        or "vardnadshavare-fklass" in filename_lower
    ):
        return "vardnadshavare-forskoleklass", "forskoleklass"
    elif (
        "vardnadshavare-grundskola" in filename_lower
        or "vardnadshavare-grundskola-ak-1-9" in filename_lower
    ):
        return "vardnadshavare-grundskola", "grundskola"
    elif (
        "vardnadshavare-anpassad-grundskola" in filename_lower
        or "vardnadshavare-anp-grundskola" in filename_lower
        or "vardnadshavare-grundsarskola" in filename_lower
    ):
        return "vardnadshavare-anpassad-grundskola", "anpassad-grundskola"
    elif "pedagogisk-personal-forskola" in filename_lower:
        return "pedagogisk-personal-forskola", "forskola"
    elif "vardnadshavare-forskola" in filename_lower:
        return "vardnadshavare-forskola", "forskola"

    # Default
    return "unknown", None


def parse_year_from_path(file_path: Path) -> int:
    """Extract year from file path.

    Args:
        file_path: Path to the Excel file

    Returns:
        Year as integer
    """
    # Look for year in path components
    path_str = str(file_path)

    # Try to find 4-digit year (2015-2030 range)
    year_match = re.search(r"/(20[12]\d)/", path_str)
    if year_match:
        return int(year_match.group(1))

    # Try filename
    year_match = re.search(r"(\d{4})", file_path.name)
    if year_match:
        year = int(year_match.group(1))
        if 2010 <= year <= 2030:
            return year

    # Try vt/ht pattern
    term_match = re.search(r"(vt|ht)-?(\d{4})", path_str.lower())
    if term_match:
        return int(term_match.group(2))

    # Default to current year
    return 2025


def parse_term_from_path(file_path: Path) -> Optional[str]:
    """Extract term (semester) from file path.

    Args:
        file_path: Path to the Excel file

    Returns:
        'vt' (spring), 'ht' (fall), or None
    """
    path_str = str(file_path).lower()

    if "vt-" in path_str or "vt_" in path_str or "/vt" in path_str:
        return "vt"
    elif "ht-" in path_str or "ht_" in path_str or "/ht" in path_str:
        return "ht"

    return None


def safe_float(value) -> Optional[float]:
    """Safely convert a value to float.

    Args:
        value: Value to convert

    Returns:
        Float value or None
    """
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value) if value != "-" else None
    if isinstance(value, str):
        if value == "-" or value.strip() == "":
            return None
        try:
            return float(value.replace(",", "."))
        except ValueError:
            return None
    return None


def safe_int(value) -> Optional[int]:
    """Safely convert a value to int.

    Args:
        value: Value to convert

    Returns:
        Integer value or None
    """
    if value is None:
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value)
    if isinstance(value, str):
        if value == "-" or value.strip() == "":
            return None
        try:
            return int(float(value.replace(",", ".")))
        except ValueError:
            return None
    return None


def safe_str(value) -> Optional[str]:
    """Safely convert a value to string.

    Args:
        value: Value to convert

    Returns:
        String value or None
    """
    if value is None:
        return None
    if isinstance(value, str):
        return value.strip() if value.strip() else None
    return str(value)


def parse_skolenkaten_excel(file_path: Path, limit: Optional[int] = None) -> list[SkolenkatResult]:
    """Parse a Skolenkäten Excel file and extract survey results.

    Args:
        file_path: Path to the Excel file
        limit: Maximum number of results to return (None for all)

    Returns:
        List of SkolenkatResult objects
    """
    try:
        import openpyxl
    except ImportError:
        logger.error("openpyxl not installed. Install with: pip install openpyxl")
        return []

    results = []

    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    except Exception as e:
        logger.error(f"Failed to open Excel file {file_path}: {e}")
        return []

    # Get the first sheet (main results)
    sheet_names = wb.sheetnames
    if not sheet_names:
        logger.warning(f"No sheets found in {file_path}")
        return []

    ws = wb[sheet_names[0]]

    # Extract metadata from file path
    year = parse_year_from_path(file_path)
    term = parse_term_from_path(file_path)
    respondent_type, skolform = parse_respondent_type(file_path.name)

    # Find the data start row (skip header rows)
    # Usually starts at row 4 (0-indexed row 3)
    data_start_row = 4

    row_count = 0
    for row_idx, row in enumerate(ws.iter_rows(min_row=data_start_row, values_only=True)):
        # Skip empty rows or header-like rows
        if not row or row[COL_SKOLENHET] is None:
            continue

        # Skip aggregate rows (e.g., "Samtliga deltagande skolor")
        huvudman = safe_str(row[COL_HUVUDMAN])
        if huvudman and "samtliga" in huvudman.lower():
            continue

        skolenhetskod = safe_str(row[COL_SKOLENHETSKOD])
        if not skolenhetskod:
            continue

        try:
            result = SkolenkatResult(
                org_nummer=safe_str(row[COL_ORG_NUMMER]),
                huvudman=huvudman or "Okänd",
                kommun=safe_str(row[COL_KOMMUN]),
                skolenhetskod=skolenhetskod,
                skolenhet=safe_str(row[COL_SKOLENHET]) or "Okänd",
                antal_i_gruppen=safe_int(row[COL_ANTAL_I_GRUPPEN]),
                antal_svar=safe_int(row[COL_ANTAL_SVAR]),
                svarsfrekvens=safe_float(row[COL_SVARSFREKVENS]),
                year=year,
                term=term,
                respondent_type=respondent_type,
                skolform=skolform,
                index_information=safe_float(row[INDEX_POSITIONS["information"]])
                if len(row) > INDEX_POSITIONS["information"]
                else None,
                index_stimulans=safe_float(row[INDEX_POSITIONS["stimulans"]])
                if len(row) > INDEX_POSITIONS["stimulans"]
                else None,
                index_stod=safe_float(row[INDEX_POSITIONS["stod"]])
                if len(row) > INDEX_POSITIONS["stod"]
                else None,
                index_kritiskt_tankande=safe_float(row[INDEX_POSITIONS["kritiskt_tankande"]])
                if len(row) > INDEX_POSITIONS["kritiskt_tankande"]
                else None,
                index_bemotande_larare=safe_float(row[INDEX_POSITIONS["bemotande_larare"]])
                if len(row) > INDEX_POSITIONS["bemotande_larare"]
                else None,
                index_bemotande_elever=safe_float(row[INDEX_POSITIONS["bemotande_elever"]])
                if len(row) > INDEX_POSITIONS["bemotande_elever"]
                else None,
                index_inflytande=safe_float(row[INDEX_POSITIONS["inflytande"]])
                if len(row) > INDEX_POSITIONS["inflytande"]
                else None,
                index_studiero=safe_float(row[INDEX_POSITIONS["studiero"]])
                if len(row) > INDEX_POSITIONS["studiero"]
                else None,
                index_trygghet=safe_float(row[INDEX_POSITIONS["trygghet"]])
                if len(row) > INDEX_POSITIONS["trygghet"]
                else None,
                index_forhindra_krankningar=safe_float(
                    row[INDEX_POSITIONS["forhindra_krankningar"]]
                )
                if len(row) > INDEX_POSITIONS["forhindra_krankningar"]
                else None,
                index_elevhalsa=safe_float(row[INDEX_POSITIONS["elevhalsa"]])
                if len(row) > INDEX_POSITIONS["elevhalsa"]
                else None,
                index_nojdhet=safe_float(row[INDEX_POSITIONS["nojdhet"]])
                if len(row) > INDEX_POSITIONS["nojdhet"]
                else None,
            )
            results.append(result)
            row_count += 1

            if limit and row_count >= limit:
                break

        except Exception as e:
            logger.debug(f"Skipping row due to error: {e}")
            continue

    logger.info(f"Parsed {len(results)} results from {file_path.name}")
    return results


def create_summary(results: list[SkolenkatResult]) -> Optional[SkolenkatSummary]:
    """Create a summary from a list of Skolenkäten results.

    Args:
        results: List of SkolenkatResult objects

    Returns:
        SkolenkatSummary object or None if no results
    """
    if not results:
        return None

    first = results[0]

    # Calculate totals
    total_responses = sum(r.antal_svar or 0 for r in results)
    response_rates = [r.svarsfrekvens for r in results if r.svarsfrekvens is not None]
    avg_response_rate = sum(response_rates) / len(response_rates) if response_rates else None

    # Calculate national averages for each index
    def calc_avg(values: list[Optional[float]]) -> Optional[float]:
        valid = [v for v in values if v is not None]
        return sum(valid) / len(valid) if valid else None

    return SkolenkatSummary(
        year=first.year,
        term=first.term,
        respondent_type=first.respondent_type,
        total_schools=len(results),
        total_responses=total_responses,
        average_response_rate=avg_response_rate,
        national_index_information=calc_avg([r.index_information for r in results]),
        national_index_stimulans=calc_avg([r.index_stimulans for r in results]),
        national_index_stod=calc_avg([r.index_stod for r in results]),
        national_index_kritiskt_tankande=calc_avg([r.index_kritiskt_tankande for r in results]),
        national_index_bemotande_larare=calc_avg([r.index_bemotande_larare for r in results]),
        national_index_bemotande_elever=calc_avg([r.index_bemotande_elever for r in results]),
        national_index_inflytande=calc_avg([r.index_inflytande for r in results]),
        national_index_studiero=calc_avg([r.index_studiero for r in results]),
        national_index_trygghet=calc_avg([r.index_trygghet for r in results]),
        national_index_forhindra_krankningar=calc_avg(
            [r.index_forhindra_krankningar for r in results]
        ),
        national_index_elevhalsa=calc_avg([r.index_elevhalsa for r in results]),
        national_index_nojdhet=calc_avg([r.index_nojdhet for r in results]),
    )


def discover_skolenkaten_files(base_path: Path) -> list[Path]:
    """Discover all Skolenkäten Excel files in a directory tree.

    Args:
        base_path: Base directory to search

    Returns:
        List of paths to Skolenkäten Excel files
    """
    patterns = [
        "**/statistik-skolenkaten/**/*.xlsx",
        "**/skolenkaten/**/*.xlsx",
        "**/*.xlsx",  # Also search directly in the given directory
    ]

    files = []
    for pattern in patterns:
        files.extend(base_path.glob(pattern))

    # Filter out temp files and duplicates
    seen = set()
    result = []
    for f in files:
        if f.name.startswith("~$"):
            continue
        if f.name in seen:
            continue
        seen.add(f.name)
        result.append(f)

    return sorted(result, key=lambda p: (parse_year_from_path(p), p.name), reverse=True)


def search_schools_in_results(
    results: list[SkolenkatResult],
    query: str,
    kommun: Optional[str] = None,
    huvudman: Optional[str] = None,
) -> list[SkolenkatResult]:
    """Search for schools in Skolenkäten results.

    Args:
        results: List of results to search
        query: Search query (matches school name)
        kommun: Optional municipality filter
        huvudman: Optional operator filter

    Returns:
        Filtered list of results
    """
    query_lower = query.lower()

    filtered = []
    for r in results:
        # Check school name match
        if query_lower not in r.skolenhet.lower():
            continue

        # Check kommun filter
        if kommun and r.kommun:
            if kommun.lower() not in r.kommun.lower():
                continue

        # Check huvudman filter
        if huvudman and r.huvudman:
            if huvudman.lower() not in r.huvudman.lower():
                continue

        filtered.append(r)

    return filtered
