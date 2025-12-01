"""Parser for Tillståndsbeslut Excel files.

This module provides functionality to parse permit decisions (tillståndsbeslut)
for independent schools (fristående skolor) from Excel files downloaded from
Skolinspektionen's website.
"""

import logging
import re
from pathlib import Path
from typing import Optional

from .models import TillstandBeslut, TillstandSummary

logger = logging.getLogger(__name__)

# Column indices in the Excel file (0-indexed)
COL_ARENDENUMMER = 1
COL_KOMMUN = 2
COL_SKOLA = 3
COL_SOKANDE = 4
COL_SKOLFORM = 5
COL_ANSOKNINGSTYP = 6
COL_BESLUTSTYP = 7
# Grundskola grade columns
COL_AK1 = 8
COL_AK2 = 9
COL_AK3 = 10
COL_AK4 = 11
COL_AK5 = 12
COL_AK6 = 13
COL_AK7 = 14
COL_AK8 = 15
COL_AK9 = 16
COL_FORSKOLEKLASS = 17
COL_FRITIDSHEM = 18

# Gymnasieskola program columns start at 19
GYMNASIE_PROGRAM_COLS = {
    19: "Samhällsvetenskapsprogrammet",
    20: "Naturvetenskapsprogrammet",
    21: "Ekonomiprogrammet",
    22: "Estetiska programmet",
    23: "El- och energiprogrammet",
    24: "Fordons- och transportprogrammet",
    25: "Vård- och omsorgsprogrammet",
    26: "Barn- och fritidsprogrammet",
    27: "Naturbruksprogrammet",
    28: "Hotell- och turismprogrammet",
    29: "Riksrekryterande utbildningar",
    30: "Nationellt godkända idrottsutbildningar",
    31: "Särskild variant estetisk",
    32: "Särskilda varianter",
}


def parse_year_from_path(file_path: Path) -> int:
    """Extract decision year from file path.

    Args:
        file_path: Path to the Excel file

    Returns:
        Year as integer
    """
    path_str = str(file_path)

    # Try to find year in path (e.g., "2023-skolstart-2024-25")
    year_match = re.search(r"/(20\d{2})-skolstart", path_str)
    if year_match:
        return int(year_match.group(1))

    # Try filename patterns
    year_match = re.search(r"tillstandsbeslut-(\d{4})", file_path.name.lower())
    if year_match:
        return int(year_match.group(1))

    # Try direct year in filename
    year_match = re.search(r"(20\d{2})", file_path.name)
    if year_match:
        return int(year_match.group(1))

    # Default to current year
    return 2025


def parse_skolstart_from_path(file_path: Path) -> str:
    """Extract school start year from file path.

    Args:
        file_path: Path to the Excel file

    Returns:
        School year string (e.g., "2024-25")
    """
    path_str = str(file_path)

    # Try to find skolstart pattern (e.g., "2023-skolstart-2024-25")
    skolstart_match = re.search(r"skolstart-(\d{4}-\d{2})", path_str)
    if skolstart_match:
        return skolstart_match.group(1)

    # Calculate from decision year
    year = parse_year_from_path(file_path)
    next_year = year + 1
    return f"{next_year}-{str(next_year + 1)[-2:]}"


def safe_str(value) -> Optional[str]:
    """Safely convert a value to string.

    Args:
        value: Value to convert

    Returns:
        String value or None
    """
    if value is None:
        return None
    if isinstance(value, str):
        return value.strip() if value.strip() else None
    return str(value)


def parse_tillstand_excel(
    file_path: Path, limit: Optional[int] = None
) -> list[TillstandBeslut]:
    """Parse a Tillståndsbeslut Excel file and extract permit decisions.

    Args:
        file_path: Path to the Excel file
        limit: Maximum number of results to return (None for all)

    Returns:
        List of TillstandBeslut objects
    """
    try:
        import openpyxl
    except ImportError:
        logger.error("openpyxl not installed. Install with: pip install openpyxl")
        return []

    results = []

    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    except Exception as e:
        logger.error(f"Failed to open Excel file {file_path}: {e}")
        return []

    # Look for "Skola för skola" sheet
    target_sheet = None
    for name in wb.sheetnames:
        if "skola för skola" in name.lower():
            target_sheet = name
            break

    if not target_sheet:
        logger.warning(f"No 'Skola för skola' sheet found in {file_path}")
        return []

    ws = wb[target_sheet]

    # Extract metadata from file path
    year = parse_year_from_path(file_path)
    skolstart_lasar = parse_skolstart_from_path(file_path)

    # Find the data start row (look for "Ärendenummer" header)
    data_start_row = None
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=30, values_only=True), 1):
        if row and len(row) > COL_ARENDENUMMER:
            cell = row[COL_ARENDENUMMER]
            if cell and "rendenummer" in str(cell):
                data_start_row = row_idx + 1
                break

    if data_start_row is None:
        logger.warning(f"Could not find header row in {file_path}")
        return []

    row_count = 0
    for row in ws.iter_rows(min_row=data_start_row, values_only=True):
        # Skip empty rows
        if not row or len(row) <= COL_BESLUTSTYP:
            continue

        arendenummer = safe_str(row[COL_ARENDENUMMER])
        if not arendenummer or not arendenummer.startswith("SI"):
            continue

        kommun = safe_str(row[COL_KOMMUN])
        skola = safe_str(row[COL_SKOLA])
        sokande = safe_str(row[COL_SOKANDE])
        skolform = safe_str(row[COL_SKOLFORM])
        ansokningstyp = safe_str(row[COL_ANSOKNINGSTYP])
        beslutstyp = safe_str(row[COL_BESLUTSTYP])

        # Skip if missing required fields
        if not all([kommun, skola, sokande, skolform, beslutstyp]):
            continue

        # Parse grade-level decisions
        beslut_ak1 = safe_str(row[COL_AK1]) if len(row) > COL_AK1 else None
        beslut_ak2 = safe_str(row[COL_AK2]) if len(row) > COL_AK2 else None
        beslut_ak3 = safe_str(row[COL_AK3]) if len(row) > COL_AK3 else None
        beslut_ak4 = safe_str(row[COL_AK4]) if len(row) > COL_AK4 else None
        beslut_ak5 = safe_str(row[COL_AK5]) if len(row) > COL_AK5 else None
        beslut_ak6 = safe_str(row[COL_AK6]) if len(row) > COL_AK6 else None
        beslut_ak7 = safe_str(row[COL_AK7]) if len(row) > COL_AK7 else None
        beslut_ak8 = safe_str(row[COL_AK8]) if len(row) > COL_AK8 else None
        beslut_ak9 = safe_str(row[COL_AK9]) if len(row) > COL_AK9 else None
        beslut_forskoleklass = safe_str(row[COL_FORSKOLEKLASS]) if len(row) > COL_FORSKOLEKLASS else None
        beslut_fritidshem = safe_str(row[COL_FRITIDSHEM]) if len(row) > COL_FRITIDSHEM else None

        # Parse gymnasieskola program decisions
        gymnasie_programs = None
        if skolform and "gymnasi" in skolform.lower():
            gymnasie_programs = {}
            for col_idx, program_name in GYMNASIE_PROGRAM_COLS.items():
                if len(row) > col_idx:
                    decision = safe_str(row[col_idx])
                    if decision:
                        gymnasie_programs[program_name] = decision

            if not gymnasie_programs:
                gymnasie_programs = None

        try:
            result = TillstandBeslut(
                year=year,
                skolstart_lasar=skolstart_lasar,
                arendenummer=arendenummer,
                kommun=kommun,
                skola=skola,
                sokande=sokande,
                skolform=skolform,
                ansokningstyp=ansokningstyp or "Okänd",
                beslutstyp=beslutstyp,
                beslut_ak1=beslut_ak1,
                beslut_ak2=beslut_ak2,
                beslut_ak3=beslut_ak3,
                beslut_ak4=beslut_ak4,
                beslut_ak5=beslut_ak5,
                beslut_ak6=beslut_ak6,
                beslut_ak7=beslut_ak7,
                beslut_ak8=beslut_ak8,
                beslut_ak9=beslut_ak9,
                beslut_forskoleklass=beslut_forskoleklass,
                beslut_fritidshem=beslut_fritidshem,
                gymnasie_programs=gymnasie_programs,
            )
            results.append(result)
            row_count += 1

            if limit and row_count >= limit:
                break

        except Exception as e:
            logger.debug(f"Skipping row due to error: {e}")
            continue

    logger.info(f"Parsed {len(results)} decisions from {file_path.name}")
    return results


def create_summary(results: list[TillstandBeslut]) -> Optional[TillstandSummary]:
    """Create a summary from a list of tillståndsbeslut.

    Args:
        results: List of TillstandBeslut objects

    Returns:
        TillstandSummary object or None if no results
    """
    if not results:
        return None

    first = results[0]

    # Count by decision type
    godkannanden = sum(1 for r in results if "godkännande" in r.beslutstyp.lower())
    avslag = sum(1 for r in results if "avslag" in r.beslutstyp.lower())
    avskrivningar = sum(1 for r in results if "avskriv" in r.beslutstyp.lower())

    # Count by application type
    nyetableringar = [r for r in results if "nyetablering" in r.ansokningstyp.lower()]
    utokningar = [r for r in results if "utökning" in r.ansokningstyp.lower()]

    nyetableringar_godkanda = sum(
        1 for r in nyetableringar if "godkännande" in r.beslutstyp.lower()
    )
    utokningar_godkanda = sum(
        1 for r in utokningar if "godkännande" in r.beslutstyp.lower()
    )

    # Count by school form
    by_skolform: dict[str, dict[str, int]] = {}
    for r in results:
        sf = r.skolform
        if sf not in by_skolform:
            by_skolform[sf] = {"total": 0, "godkannanden": 0, "avslag": 0}
        by_skolform[sf]["total"] += 1
        if "godkännande" in r.beslutstyp.lower():
            by_skolform[sf]["godkannanden"] += 1
        elif "avslag" in r.beslutstyp.lower():
            by_skolform[sf]["avslag"] += 1

    return TillstandSummary(
        year=first.year,
        skolstart_lasar=first.skolstart_lasar,
        total_decisions=len(results),
        godkannanden=godkannanden,
        avslag=avslag,
        avskrivningar=avskrivningar,
        nyetableringar_total=len(nyetableringar),
        nyetableringar_godkanda=nyetableringar_godkanda,
        utokningar_total=len(utokningar),
        utokningar_godkanda=utokningar_godkanda,
        by_skolform=by_skolform,
    )


def discover_tillstand_files(base_path: Path) -> list[Path]:
    """Discover all Tillståndsbeslut Excel files in a directory tree.

    Args:
        base_path: Base directory to search

    Returns:
        List of paths to Tillståndsbeslut Excel files
    """
    patterns = [
        "**/statistik-tillstand/**/*.xlsx",
        "**/tillstand/**/*.xlsx",
        "**/*.xlsx",
    ]

    files = []
    for pattern in patterns:
        files.extend(base_path.glob(pattern))

    # Filter to only include decision files (not just application statistics)
    result = []
    seen = set()
    for f in files:
        if f.name.startswith("~$"):
            continue
        if f.name in seen:
            continue
        # Only include files that look like decision files
        name_lower = f.name.lower()
        if "tillstandsbeslut" in name_lower or "beslut" in name_lower:
            seen.add(f.name)
            result.append(f)

    return sorted(result, key=lambda p: (parse_year_from_path(p), p.name), reverse=True)


def search_tillstand(
    results: list[TillstandBeslut],
    query: Optional[str] = None,
    kommun: Optional[str] = None,
    skolform: Optional[str] = None,
    beslutstyp: Optional[str] = None,
    ansokningstyp: Optional[str] = None,
) -> list[TillstandBeslut]:
    """Search tillståndsbeslut with various filters.

    Args:
        results: List of decisions to search
        query: Free text search (school name, applicant)
        kommun: Municipality filter
        skolform: School form filter
        beslutstyp: Decision type filter (Godkännande, Avslag, etc.)
        ansokningstyp: Application type filter (Nyetablering, Utökning)

    Returns:
        Filtered list of decisions
    """
    filtered = results

    if query:
        query_lower = query.lower()
        filtered = [
            r for r in filtered
            if query_lower in r.skola.lower() or query_lower in r.sokande.lower()
        ]

    if kommun:
        kommun_lower = kommun.lower()
        filtered = [r for r in filtered if kommun_lower in r.kommun.lower()]

    if skolform:
        skolform_lower = skolform.lower()
        filtered = [r for r in filtered if skolform_lower in r.skolform.lower()]

    if beslutstyp:
        beslutstyp_lower = beslutstyp.lower()
        filtered = [r for r in filtered if beslutstyp_lower in r.beslutstyp.lower()]

    if ansokningstyp:
        ansokningstyp_lower = ansokningstyp.lower()
        filtered = [r for r in filtered if ansokningstyp_lower in r.ansokningstyp.lower()]

    return filtered
