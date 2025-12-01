"""Parser for Tillsyn statistics Excel files.

This module provides functionality to parse supervision statistics
from Excel files downloaded from Skolinspektionen's website, including:
- Viten (fines) statistics
- TUI (Tillsyn utifran individärenden) / BEO statistics
- Planerad tillsyn statistics
"""

import logging
import re
from pathlib import Path
from typing import Optional

from .models import (
    ViteStatistik,
    TUIStatistik,
    PlaneradTillsynStatistik,
    TillsynStatistikSummary,
)

logger = logging.getLogger(__name__)


def safe_int(value) -> int:
    """Safely convert a value to int, defaulting to 0."""
    if value is None:
        return 0
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value)
    if isinstance(value, str):
        if value == "-" or value.strip() == "":
            return 0
        try:
            return int(float(value.replace(",", ".")))
        except ValueError:
            return 0
    return 0


def safe_float(value) -> Optional[float]:
    """Safely convert a value to float."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        if value == "-" or value.strip() == "":
            return None
        try:
            return float(value.replace(",", "."))
        except ValueError:
            return None
    return None


def parse_viten_excel(file_path: Path) -> list[ViteStatistik]:
    """Parse Viten (fines) statistics Excel file.

    Args:
        file_path: Path to the Excel file

    Returns:
        List of ViteStatistik objects, one per year
    """
    try:
        import openpyxl
    except ImportError:
        logger.error("openpyxl not installed. Install with: pip install openpyxl")
        return []

    results = []

    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    except Exception as e:
        logger.error(f"Failed to open Excel file {file_path}: {e}")
        return []

    # Look for Tabeller sheet
    if "Tabeller" not in wb.sheetnames:
        logger.warning(f"No 'Tabeller' sheet found in {file_path}")
        return []

    ws = wb["Tabeller"]

    # Parse Table 1: Antal beslut om vitesförelägganden
    # Parse Table 2: Antal ansökningar om utdömande av vite

    beslut_data = {}  # year -> {total, enskild, offentlig}
    ansokningar_data = {}  # year -> {total, enskild, offentlig}

    current_table = None
    for row in ws.iter_rows(min_row=1, max_row=100, values_only=True):
        if not row or row[1] is None:
            continue

        cell_value = row[1]
        cell = str(cell_value)

        # Detect which table we're in
        if "Tabell 1" in cell and "vite" in cell.lower():
            current_table = "beslut"
            continue
        elif "Tabell 2" in cell and "ansökningar" in cell.lower():
            current_table = "ansokningar"
            continue

        # Parse year rows - years can be integers or strings like "2017**"
        year = None
        if isinstance(cell_value, int) and 2000 <= cell_value <= 2030:
            year = cell_value
        elif isinstance(cell_value, str) and re.match(r"^\d{4}\*{0,2}$", cell.strip()):
            year = int(cell.strip().replace("*", ""))

        if current_table and year:
            total = safe_int(row[2])
            enskild = safe_int(row[3])
            offentlig = safe_int(row[4])

            if current_table == "beslut":
                beslut_data[year] = {
                    "total": total,
                    "enskild": enskild,
                    "offentlig": offentlig,
                }
            elif current_table == "ansokningar":
                ansokningar_data[year] = {
                    "total": total,
                    "enskild": enskild,
                    "offentlig": offentlig,
                }

    # Combine data into results
    all_years = set(beslut_data.keys()) | set(ansokningar_data.keys())
    for year in sorted(all_years, reverse=True):
        beslut = beslut_data.get(year, {})
        ansok = ansokningar_data.get(year, {})

        results.append(
            ViteStatistik(
                year=year,
                beslut_totalt=beslut.get("total", 0),
                beslut_enskild=beslut.get("enskild", 0),
                beslut_offentlig=beslut.get("offentlig", 0),
                ansokningar_totalt=ansok.get("total", 0),
                ansokningar_enskild=ansok.get("enskild", 0),
                ansokningar_offentlig=ansok.get("offentlig", 0),
            )
        )

    logger.info(f"Parsed {len(results)} years of Viten statistics from {file_path.name}")
    return results


def parse_tui_excel(file_path: Path) -> Optional[TUIStatistik]:
    """Parse TUI (individual case) statistics Excel file.

    Args:
        file_path: Path to the Excel file

    Returns:
        TUIStatistik object for the year
    """
    try:
        import openpyxl
    except ImportError:
        logger.error("openpyxl not installed. Install with: pip install openpyxl")
        return None

    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    except Exception as e:
        logger.error(f"Failed to open Excel file {file_path}: {e}")
        return None

    # Extract year from filename or path
    year_match = re.search(r"(\d{4})", file_path.name)
    if not year_match:
        year_match = re.search(r"(\d{4})", str(file_path))
    year = int(year_match.group(1)) if year_match else 2024

    # Look for Tabeller sheet
    if "Tabeller" not in wb.sheetnames:
        logger.warning(f"No 'Tabeller' sheet found in {file_path}")
        return None

    ws = wb["Tabeller"]

    result = TUIStatistik(year=year)

    current_table = None
    for row in ws.iter_rows(min_row=1, max_row=100, values_only=True):
        if not row or row[1] is None:
            continue

        cell = str(row[1])

        # Detect which table we're in
        if "Tabell 1" in cell:
            current_table = "table1"
            continue
        elif "Tabell 2" in cell:
            current_table = "table2"
            continue
        elif "Tabell 3" in cell:
            current_table = "table3"
            continue

        # Parse Table 1: Antal beslut och brister
        if current_table == "table1":
            if "Antal beslut totalt" in cell:
                result.beslut_totalt = safe_int(row[2])
                result.beslut_flickor = safe_int(row[3])
                result.beslut_pojkar = safe_int(row[4])
                result.beslut_ovriga = safe_int(row[5])
                result.beslut_enskild = safe_int(row[6])
                result.beslut_offentlig = safe_int(row[10])
            elif "antal beslut med brist" in cell.lower():
                result.beslut_med_brist = safe_int(row[2])
                result.beslut_enskild_med_brist = safe_int(row[6])
                result.beslut_offentlig_med_brist = safe_int(row[10])
            elif "andel beslut med brist" in cell.lower():
                result.andel_med_brist = safe_float(row[2])

        # Parse Table 2: Antal beslut per verksamhetsform
        elif current_table == "table2":
            if cell in [
                "Förskola",
                "Grundskola",
                "Anpassad grundskola",
                "Gymnasieskola",
                "Anpassad gymnasieskola",
                "Komvux",
                "SFI",
            ]:
                result.by_skolform[cell] = safe_int(row[2])

        # Parse Table 3: Antal beslut med brist per bedömningsområde
        elif current_table == "table3":
            if "Kränkande behandling" in cell and "varav" not in cell.lower():
                result.brister_krankande_behandling = safe_int(row[2])
            elif "varav elev-elev" in cell.lower():
                result.brister_elev_elev = safe_int(row[2])
            elif "varav personal-elev" in cell.lower():
                result.brister_personal_elev = safe_int(row[2])
            elif "Stöd" in cell and "Särskilt stöd" in cell:
                result.brister_stod = safe_int(row[2])
            elif "Undervisning" in cell:
                result.brister_undervisning = safe_int(row[2])
            elif "Övriga" in cell or "Ytterligare" in cell:
                result.brister_ovriga = safe_int(row[2])

    logger.info(f"Parsed TUI statistics for year {year} from {file_path.name}")
    return result


def parse_planerad_tillsyn_excel(file_path: Path) -> Optional[PlaneradTillsynStatistik]:
    """Parse Planerad Tillsyn statistics Excel file.

    Args:
        file_path: Path to the Excel file

    Returns:
        PlaneradTillsynStatistik object for the year
    """
    try:
        import openpyxl
    except ImportError:
        logger.error("openpyxl not installed. Install with: pip install openpyxl")
        return None

    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    except Exception as e:
        logger.error(f"Failed to open Excel file {file_path}: {e}")
        return None

    # Extract year from filename or path
    year_match = re.search(r"(\d{4})", file_path.name)
    if not year_match:
        year_match = re.search(r"(\d{4})", str(file_path))
    year = int(year_match.group(1)) if year_match else 2024

    # Look for Tabeller sheet
    if "Tabeller" not in wb.sheetnames:
        logger.warning(f"No 'Tabeller' sheet found in {file_path}")
        return None

    ws = wb["Tabeller"]

    result = PlaneradTillsynStatistik(year=year)

    current_table = None
    for row in ws.iter_rows(min_row=1, max_row=100, values_only=True):
        if not row or row[1] is None:
            continue

        cell = str(row[1])

        # Detect which table we're in
        if "Tabell 1" in cell:
            current_table = "table1"
            continue
        elif "Tabell 2" in cell:
            current_table = "table2"
            continue

        # Parse Table 1: Antal beslut och brister totalt
        if current_table == "table1":
            if "Antal beslut totalt" in cell:
                result.beslut_totalt = safe_int(row[2])
                result.beslut_enskild = safe_int(row[3])
                result.beslut_offentlig = safe_int(row[4])
            elif "antal beslut med brist" in cell.lower():
                result.beslut_med_brist = safe_int(row[2])
                result.beslut_enskild_med_brist = safe_int(row[3])
                result.beslut_offentlig_med_brist = safe_int(row[4])
            elif "andel beslut med brist" in cell.lower():
                result.andel_med_brist = safe_float(row[2])

        # Parse Table 2: Antal beslut per skolform
        elif current_table == "table2":
            if cell in [
                "Grundskola",
                "Anpassad grundskola",
                "Gymnasieskola",
                "Anpassad gymnasieskola",
                "Komvux",
                "Sameskola",
                "Specialskola",
            ]:
                total = safe_int(row[2])
                med_brist = safe_int(row[3])
                result.by_skolform[cell] = {"total": total, "med_brist": med_brist}

    logger.info(f"Parsed Planerad Tillsyn statistics for year {year} from {file_path.name}")
    return result


def discover_tillsyn_files(base_path: Path) -> dict[str, list[Path]]:
    """Discover all Tillsyn statistics Excel files.

    Args:
        base_path: Base directory to search

    Returns:
        Dict mapping category to list of file paths
    """
    files = {
        "viten": [],
        "tui": [],
        "planerad_tillsyn": [],
    }

    # Search patterns - updated to match actual Skolinspektionen file structure
    patterns = {
        "viten": [
            "**/statistik-viten/**/*.xlsx",
            "**/viten/**/*.xlsx",
            "**/*vite*.xlsx",
        ],
        "tui": [
            # RT individ files (riktad tillsyn individ = same as TUI/BEO)
            "**/rt-*-individ/**/*.xlsx",
            "**/rt-individ-*/**/*.xlsx",
            "**/riktad-tillsyn-individ*.xlsx",
            "**/statistik-riktad-tillsyn-individ*.xlsx",
            # Direct TUI files
            "**/statistik-tui/**/*.xlsx",
            "**/tui-*/**/*.xlsx",
        ],
        "planerad_tillsyn": [
            "**/planerad-tillsyn/**/*.xlsx",
            "**/pt-*/**/*.xlsx",
            "**/statistik-planerad-tillsyn*.xlsx",
            "**/arsstatistik-*.xlsx",
        ],
    }

    for category, pattern_list in patterns.items():
        seen = set()
        for pattern in pattern_list:
            for f in base_path.glob(pattern):
                if f.name.startswith("~$"):
                    continue
                if f.name in seen:
                    continue
                seen.add(f.name)
                files[category].append(f)

    return files


def load_all_tillsyn_statistik(base_path: Path) -> TillsynStatistikSummary:
    """Load all Tillsyn statistics from Excel files.

    Args:
        base_path: Base directory containing statistics files

    Returns:
        TillsynStatistikSummary with all parsed data
    """
    files = discover_tillsyn_files(base_path)
    summary = TillsynStatistikSummary()

    # Parse Viten files
    for f in files["viten"]:
        viten_list = parse_viten_excel(f)
        summary.viten.extend(viten_list)

    # Parse TUI files
    for f in files["tui"]:
        tui = parse_tui_excel(f)
        if tui:
            # Avoid duplicates
            if not any(t.year == tui.year for t in summary.tui):
                summary.tui.append(tui)

    # Parse Planerad Tillsyn files
    for f in files["planerad_tillsyn"]:
        pt = parse_planerad_tillsyn_excel(f)
        if pt:
            # Avoid duplicates
            if not any(p.year == pt.year for p in summary.planerad_tillsyn):
                summary.planerad_tillsyn.append(pt)

    # Sort by year (most recent first)
    summary.viten.sort(key=lambda x: x.year, reverse=True)
    summary.tui.sort(key=lambda x: x.year, reverse=True)
    summary.planerad_tillsyn.sort(key=lambda x: x.year, reverse=True)

    # Collect all years
    all_years = set()
    all_years.update(v.year for v in summary.viten)
    all_years.update(t.year for t in summary.tui)
    all_years.update(p.year for p in summary.planerad_tillsyn)
    summary.years_available = sorted(all_years, reverse=True)

    return summary


def get_viten_by_year(
    data: list[ViteStatistik], year: Optional[int] = None
) -> list[ViteStatistik]:
    """Filter Viten statistics by year.

    Args:
        data: List of ViteStatistik
        year: Optional year filter

    Returns:
        Filtered list
    """
    if year:
        return [v for v in data if v.year == year]
    return data


def get_tui_by_year(
    data: list[TUIStatistik], year: Optional[int] = None
) -> list[TUIStatistik]:
    """Filter TUI statistics by year.

    Args:
        data: List of TUIStatistik
        year: Optional year filter

    Returns:
        Filtered list
    """
    if year:
        return [t for t in data if t.year == year]
    return data
